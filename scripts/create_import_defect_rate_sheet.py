#!/usr/bin/env python3
"""Create/update the supplier import-inspection defect-rate analysis sheet."""

from __future__ import annotations

import datetime as dt
import os
from typing import Dict, List, Set

from openpyxl import load_workbook
from openpyxl.chart import DoughnutChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SRC_SHEET = "\uc218\uc785\uac80\uc0ac"
OUT_SHEET = "\ud611\ub825\uc0ac \uc218\uc785\uac80\uc0ac \ubd88\ub7c9\uc728"


def find_workbook_path() -> str:
    files = sorted(
        f
        for f in os.listdir(".")
        if f.lower().endswith(".xlsx") and not os.path.basename(f).startswith("~$")
    )
    if not files:
        raise FileNotFoundError("No .xlsx workbook found in current directory.")
    return files[0]


def to_float(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def collect_source_meta(ws) -> tuple[Set[int], Dict[str, float]]:
    years: Set[int] = set()
    supplier_qty: Dict[str, float] = {}

    seen_data = False
    empty_streak = 0

    for row in ws.iter_rows(min_row=4, max_col=7, values_only=True):
        date_v = row[1]
        supplier_v = row[2]
        inspect_v = row[5]
        defect_v = row[6]

        has_data = any(v not in (None, "") for v in (date_v, supplier_v, inspect_v, defect_v))
        if not has_data:
            if seen_data:
                empty_streak += 1
                if empty_streak >= 5000:
                    break
            continue

        seen_data = True
        empty_streak = 0

        if isinstance(date_v, dt.datetime):
            years.add(date_v.year)
        elif isinstance(date_v, dt.date):
            years.add(date_v.year)

        supplier = str(supplier_v).strip() if supplier_v not in (None, "") else ""
        if supplier:
            supplier_qty[supplier] = supplier_qty.get(supplier, 0.0) + to_float(inspect_v)

    return years, supplier_qty


def apply_table_border(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    thin = Side(style="thin", color="B7BDC8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = border


def create_sheet(path: str) -> str:
    wb = load_workbook(path)
    if SRC_SHEET not in wb.sheetnames:
        raise ValueError(f"Required sheet not found: {SRC_SHEET}")

    src = wb[SRC_SHEET]
    years, supplier_qty = collect_source_meta(src)
    default_year = max(years) if years else dt.date.today().year

    top_suppliers = sorted(supplier_qty.items(), key=lambda x: (-x[1], x[0]))
    supplier_names: List[str] = [name for name, _ in top_suppliers[:5]]

    if OUT_SHEET in wb.sheetnames:
        del wb[OUT_SHEET]
    ws = wb.create_sheet(OUT_SHEET)

    # Basic layout
    ws.freeze_panes = "B5"
    ws.column_dimensions["A"].width = 24
    for col in range(2, 15):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["P"].width = 2

    # Colors and fonts
    title_fill = PatternFill("solid", fgColor="DCEBFF")
    header_fill = PatternFill("solid", fgColor="F4D03F")
    label_fill = PatternFill("solid", fgColor="EFF2F7")
    summary_fill = PatternFill("solid", fgColor="DCE9FF")
    target_fill = PatternFill("solid", fgColor="FFE8C2")

    title_font = Font(size=14, bold=True, color="1F2D3D")
    header_font = Font(size=10, bold=True)
    normal_font = Font(size=10)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # Header area
    ws.merge_cells("A1:N1")
    ws["A1"] = "\ud611\ub825\uc0ac \uc218\uc785\uac80\uc0ac \ubd88\ub7c9\uc728 \ubd84\uc11d"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    ws["A2"] = "\uae30\uc900\uc5f0\ub3c4"
    ws["A2"].font = header_font
    ws["A2"].alignment = center
    ws["A2"].fill = label_fill

    ws["B2"] = default_year
    ws["B2"].font = Font(size=11, bold=True)
    ws["B2"].alignment = center
    ws["B2"].fill = PatternFill("solid", fgColor="FFFFFF")

    ws["D2"] = "\ubaa9\ud45c\ubd80\uc801\ud569\uc728(%)"
    ws["D2"].font = header_font
    ws["D2"].alignment = center
    ws["D2"].fill = label_fill

    ws["E2"] = 0.02
    ws["E2"].number_format = "0.000%"
    ws["E2"].font = Font(size=11, bold=True, color="A15C00")
    ws["E2"].alignment = center
    ws["E2"].fill = PatternFill("solid", fgColor="FFF5E6")

    # Table header
    ws["A4"] = "\uae30\uac04/\ud56d\ubaa9"
    ws["A4"].font = header_font
    ws["A4"].alignment = center
    ws["A4"].fill = header_fill

    ws["B4"] = '=$B$2&" TOTAL"'
    ws["B4"].font = header_font
    ws["B4"].alignment = center
    ws["B4"].fill = header_fill

    for month in range(1, 13):
        col = month + 2
        ws.cell(4, col, f"{month}\uc6d4")
        ws.cell(4, col).font = header_font
        ws.cell(4, col).alignment = center
        ws.cell(4, col).fill = header_fill

    # Summary rows
    ws["A5"] = "\uac80\uc0ac\uc218\ub7c9"
    ws["A6"] = "\ubd80\uc801\ud569\ud488 \uc218\ub7c9"
    ws["A7"] = "\uc804\uccb4 \ubd80\uc801\ud569\uc728"

    for row in (5, 6):
        ws[f"A{row}"].fill = label_fill
        ws[f"A{row}"].font = header_font
        ws[f"A{row}"].alignment = left

    ws["A7"].fill = summary_fill
    ws["A7"].font = header_font
    ws["A7"].alignment = left

    for col in range(2, 15):
        col_letter = get_column_letter(col)
        month = col - 2

        if col == 2:
            ws.cell(
                5,
                col,
                f"=SUMIFS('{SRC_SHEET}'!$F:$F,'{SRC_SHEET}'!$B:$B,\">=\"&DATE($B$2,1,1),'{SRC_SHEET}'!$B:$B,\"<\"&DATE($B$2+1,1,1))",
            )
            ws.cell(
                6,
                col,
                f"=SUMIFS('{SRC_SHEET}'!$G:$G,'{SRC_SHEET}'!$B:$B,\">=\"&DATE($B$2,1,1),'{SRC_SHEET}'!$B:$B,\"<\"&DATE($B$2+1,1,1))",
            )
        else:
            ws.cell(
                5,
                col,
                f"=SUMIFS('{SRC_SHEET}'!$F:$F,'{SRC_SHEET}'!$B:$B,\">=\"&DATE($B$2,{month},1),'{SRC_SHEET}'!$B:$B,\"<\"&EDATE(DATE($B$2,{month},1),1))",
            )
            ws.cell(
                6,
                col,
                f"=SUMIFS('{SRC_SHEET}'!$G:$G,'{SRC_SHEET}'!$B:$B,\">=\"&DATE($B$2,{month},1),'{SRC_SHEET}'!$B:$B,\"<\"&EDATE(DATE($B$2,{month},1),1))",
            )
        ws.cell(7, col, f"=IFERROR({col_letter}6/{col_letter}5,0)")

        ws.cell(5, col).number_format = "#,##0"
        ws.cell(6, col).number_format = "#,##0"
        ws.cell(7, col).number_format = "0.000%"
        ws.cell(5, col).alignment = center
        ws.cell(6, col).alignment = center
        ws.cell(7, col).alignment = center
        ws.cell(7, col).fill = summary_fill
        ws.cell(5, col).font = normal_font
        ws.cell(6, col).font = normal_font
        ws.cell(7, col).font = Font(size=10, bold=True, color="173A6A")

    # Supplier rows
    supplier_start = 8
    for idx, supplier in enumerate(supplier_names):
        row = supplier_start + idx
        ws.cell(row, 1, supplier)
        ws.cell(row, 1).font = normal_font
        ws.cell(row, 1).alignment = left
        ws.cell(row, 1).fill = label_fill

        for col in range(2, 15):
            month = col - 2
            if col == 2:
                formula = (
                    f"=IFERROR(SUMIFS('{SRC_SHEET}'!$G:$G,'{SRC_SHEET}'!$B:$B,\">=\"&DATE($B$2,1,1),"
                    f"'{SRC_SHEET}'!$B:$B,\"<\"&DATE($B$2+1,1,1),'{SRC_SHEET}'!$C:$C,$A{row})/"
                    f"SUMIFS('{SRC_SHEET}'!$F:$F,'{SRC_SHEET}'!$B:$B,\">=\"&DATE($B$2,1,1),"
                    f"'{SRC_SHEET}'!$B:$B,\"<\"&DATE($B$2+1,1,1),'{SRC_SHEET}'!$C:$C,$A{row}),0)"
                )
            else:
                formula = (
                    f"=IFERROR(SUMIFS('{SRC_SHEET}'!$G:$G,'{SRC_SHEET}'!$B:$B,\">=\"&DATE($B$2,{month},1),"
                    f"'{SRC_SHEET}'!$B:$B,\"<\"&EDATE(DATE($B$2,{month},1),1),'{SRC_SHEET}'!$C:$C,$A{row})/"
                    f"SUMIFS('{SRC_SHEET}'!$F:$F,'{SRC_SHEET}'!$B:$B,\">=\"&DATE($B$2,{month},1),"
                    f"'{SRC_SHEET}'!$B:$B,\"<\"&EDATE(DATE($B$2,{month},1),1),'{SRC_SHEET}'!$C:$C,$A{row}),0)"
                )
            ws.cell(row, col, formula)
            ws.cell(row, col).number_format = "0.000%"
            ws.cell(row, col).alignment = center
            ws.cell(row, col).font = normal_font

    target_row = supplier_start + len(supplier_names)
    ws.cell(target_row, 1, "\ubaa9\ud45c\ubd80\uc801\ud569\uc728")
    ws.cell(target_row, 1).font = Font(size=10, bold=True, color="A15C00")
    ws.cell(target_row, 1).alignment = left
    ws.cell(target_row, 1).fill = target_fill

    for col in range(2, 15):
        ws.cell(target_row, col, "=$E$2")
        ws.cell(target_row, col).number_format = "0.000%"
        ws.cell(target_row, col).alignment = center
        ws.cell(target_row, col).font = Font(size=10, bold=True, color="A15C00")
        ws.cell(target_row, col).fill = target_fill

    # Borders for main table
    apply_table_border(ws, 4, target_row, 1, 14)

    # Main line chart
    line_chart = LineChart()
    line_chart.title = "\ud611\ub825\uc0ac \uc218\uc785\uac80\uc0ac \ubd80\uc801\ud569\uc728 \ucd94\uc774"
    line_chart.y_axis.title = "\ubd80\uc801\ud569\uc728(%)"
    line_chart.x_axis.title = "\uae30\uac04"
    line_chart.height = 9
    line_chart.width = 25
    line_chart.legend.position = "b"

    categories = Reference(ws, min_col=2, max_col=14, min_row=4, max_row=4)
    line_chart.set_categories(categories)

    line_data = Reference(ws, min_col=1, max_col=14, min_row=7, max_row=target_row)
    line_chart.add_data(line_data, titles_from_data=True, from_rows=True)

    for idx, row in enumerate(range(7, target_row + 1)):
        series = line_chart.series[idx]
        series.marker.symbol = "circle"
        series.marker.size = 5

    line_colors = ["374151", "2E6BE6", "E11D48", "16A34A", "A855F7", "EA580C"]
    for idx, series in enumerate(line_chart.series):
        color = line_colors[idx] if idx < len(line_colors) else "64748B"
        series.graphicalProperties.line.solidFill = color
        series.graphicalProperties.line.width = 19000

    if line_chart.series:
        target_series = line_chart.series[-1]
        target_series.graphicalProperties.line.solidFill = "F59E0B"
        target_series.graphicalProperties.line.dashStyle = "sysDot"

    ws.add_chart(line_chart, f"A{target_row + 2}")

    # Defect type summary + doughnut chart
    defect_types = [src.cell(3, col).value for col in range(9, 18)]
    type_title_row = target_row + 15
    type_header_row = type_title_row + 1
    type_start_row = type_header_row + 1
    type_end_row = type_start_row + 8

    ws.merge_cells(f"A{type_title_row}:C{type_title_row}")
    ws.cell(type_title_row, 1, "\ubd88\ub7c9 \uc720\ud615 \ubd84\uc11d")
    ws.cell(type_title_row, 1).font = Font(size=12, bold=True)
    ws.cell(type_title_row, 1).fill = title_fill
    ws.cell(type_title_row, 1).alignment = center

    ws.cell(type_header_row, 1, "\uc720\ud615")
    ws.cell(type_header_row, 2, "\uc218\ub7c9")
    ws.cell(type_header_row, 3, "\ube44\uc728")
    for c in range(1, 4):
        ws.cell(type_header_row, c).fill = header_fill
        ws.cell(type_header_row, c).font = header_font
        ws.cell(type_header_row, c).alignment = center

    for idx, src_col in enumerate(range(9, 18)):
        row = type_start_row + idx
        src_col_letter = get_column_letter(src_col)
        type_name = defect_types[idx] if defect_types[idx] else f"\uc720\ud615{idx + 1}"

        ws.cell(row, 1, type_name)
        ws.cell(row, 1).alignment = left
        ws.cell(row, 1).font = normal_font
        ws.cell(row, 1).fill = label_fill

        ws.cell(
            row,
            2,
            f"=SUMIFS('{SRC_SHEET}'!${src_col_letter}:${src_col_letter},'{SRC_SHEET}'!$B:$B,\">=\"&DATE($B$2,1,1),'{SRC_SHEET}'!$B:$B,\"<\"&DATE($B$2+1,1,1))",
        )
        ws.cell(row, 2).number_format = "#,##0"
        ws.cell(row, 2).alignment = center
        ws.cell(row, 2).font = normal_font

        ws.cell(row, 3, f"=IFERROR(B{row}/SUM($B${type_start_row}:$B${type_end_row}),0)")
        ws.cell(row, 3).number_format = "0.00%"
        ws.cell(row, 3).alignment = center
        ws.cell(row, 3).font = normal_font

    apply_table_border(ws, type_header_row, type_end_row, 1, 3)

    donut = DoughnutChart()
    donut.title = "\ubd88\ub7c9 \uc720\ud615 \ube44\uc911"
    donut.holeSize = 45
    donut.height = 7
    donut.width = 13
    donut.legend.position = "b"
    donut_data = Reference(ws, min_col=2, max_col=2, min_row=type_start_row, max_row=type_end_row)
    donut_labels = Reference(ws, min_col=1, max_col=1, min_row=type_start_row, max_row=type_end_row)
    donut.add_data(donut_data, titles_from_data=False)
    donut.set_categories(donut_labels)
    ws.add_chart(donut, f"E{type_title_row}")

    note_row = type_end_row + 2
    ws.merge_cells(f"A{note_row}:N{note_row}")
    ws.cell(
        note_row,
        1,
        "\u203b \uc9d1\uacc4 \uae30\uc900: \uc218\uc785\uac80\uc0ac \ud0ed\uc758 B(\ub0a9\uae30 \uc77c\uc790), C(\uc5c5\uccb4\uba85), F(\uac80\uc0ac \uc218\ub7c9), G(\ubd80\uc801\ud569\ud488 \uc218\ub7c9).",
    )
    ws.cell(note_row, 1).font = Font(size=9, color="5F6B7A")
    ws.cell(note_row, 1).alignment = Alignment(horizontal="left", vertical="center")

    # Recalculate formulas when opened in Excel
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    saved_path = path
    try:
        wb.save(path)
    except PermissionError:
        root, ext = os.path.splitext(path)
        saved_path = f"{root}_\ud611\ub825\uc0ac\ubd88\ub7c9\uc728\ucd94\uac00{ext}"
        wb.save(saved_path)

    return saved_path


def main() -> None:
    path = find_workbook_path()
    saved_path = create_sheet(path)
    print(f"Updated workbook: {saved_path}")
    print(f"Created sheet: {OUT_SHEET}")


if __name__ == "__main__":
    main()
