#!/usr/bin/env python3
from __future__ import annotations

import argparse
import base64
import errno
import gzip
import hashlib
import hmac
import ipaddress
import json
import os
import secrets
import socket
import sys
import tempfile
import threading
from collections import OrderedDict
from datetime import datetime, timezone
from functools import partial
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from urllib.parse import parse_qs, quote, urlparse

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference

    OPENPYXL_AVAILABLE = True
except Exception:
    Workbook = None  # type: ignore[assignment]
    BarChart = None  # type: ignore[assignment]
    Reference = None  # type: ignore[assignment]
    OPENPYXL_AVAILABLE = False


# 첨부 파일 업로드 허용 최대 크기
MAX_UPLOAD_BYTES = 30 * 1024 * 1024
# 보고서 캐시(JSON) 허용 최대 크기
MAX_CACHE_BYTES  = 20 * 1024 * 1024
# 업로드 분석 캐시 스키마 버전 (구버전 캐시와 호환 끊기)
CACHE_SCHEMA_VERSION = 3
# 서버에서 허용할 첨부 파일 확장자 목록
ALLOWED_ATTACH_EXTENSIONS = {
    ".pdf",
    ".hwp",
    ".hwpx",
    ".doc",
    ".docx",
    ".xls",
    ".xlsx",
    ".ppt",
    ".pptx",
    ".jpg",
    ".jpeg",
    ".png",
    ".zip",
}

# 정적 텍스트 리소스 gzip 전송 설정 (LAN 초기 로딩 속도 개선)
GZIP_MIN_BYTES = 1400
GZIP_LEVEL = 5
GZIP_CACHE_MAX_ITEMS = 64
GZIP_STATIC_EXTENSIONS = {
    ".html",
    ".htm",
    ".js",
    ".css",
    ".json",
    ".txt",
    ".svg",
    ".csv",
    ".md",
}
GZIP_CONTENT_TYPES = {
    "application/javascript",
    "text/javascript",
    "application/json",
    "application/xml",
    "image/svg+xml",
}


def get_lan_ip() -> str:
    """현재 PC의 LAN IP를 추정해 반환한다. 실패 시 빈 문자열."""
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
            s.connect(("8.8.8.8", 80))
            return s.getsockname()[0]
    except Exception:
        return ""


def print_qr_code(url: str) -> None:
    """qrcode 패키지가 있으면 콘솔에 QR코드를 출력한다."""
    try:
        import qrcode  # type: ignore[import]
        qr = qrcode.QRCode(border=1)
        qr.add_data(url)
        qr.make(fit=True)
        print()
        print("  [QR] Scan with mobile camera to open the report URL.")
        qr.print_ascii(invert=True)
    except ImportError:
        print()
        print("  [QR] Optional: run 'pip install qrcode' to print QR code in console.")
        print()
    except Exception:
        pass


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def sanitize_filename(name: str) -> str:
    src = str(name or "file").strip()
    for ch in '\\/:*?"<>|':
        src = src.replace(ch, "_")
    src = " ".join(src.split())
    return src or "file"


def make_storage_filename(row_key: str, original_name: str) -> str:
    safe_key = quote(str(row_key or "sw_attach"), safe="").replace("%", "_")
    safe_name = sanitize_filename(original_name or "file")
    return f"{safe_key}__{safe_name}"


def default_manifest() -> dict:
    return {"version": 1, "updatedAt": "", "items": {}}


def normalize_manifest(data: object) -> dict:
    if not isinstance(data, dict):
        return default_manifest()
    items = data.get("items")
    if not isinstance(items, dict):
        items = {}
    return {
        "version": 1,
        "updatedAt": str(data.get("updatedAt") or ""),
        "items": items,
    }


def load_manifest(path: Path) -> dict:
    if not path.exists():
        return default_manifest()
    try:
        parsed = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default_manifest()
    return normalize_manifest(parsed)


def _atomic_write_json(path: Path, payload: object) -> None:
    # tmp 파일에 먼저 쓰고 교체해, 저장 중단 시 파일 손상을 줄인다.
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_name = ""
    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            dir=str(path.parent),
            prefix=f".{path.name}.",
            suffix=".tmp",
            delete=False,
        ) as tmp:
            json.dump(payload, tmp, ensure_ascii=False, indent=2)
            tmp.flush()
            os.fsync(tmp.fileno())
            tmp_name = tmp.name
        os.replace(tmp_name, path)
    except Exception:
        if tmp_name:
            try:
                os.remove(tmp_name)
            except OSError:
                pass
        raise


def save_manifest(path: Path, data: dict) -> None:
    normalized = normalize_manifest(data)
    _atomic_write_json(path, normalized)


def to_int(value: object, default: int = 0) -> int:
    try:
        return int(value)  # type: ignore[arg-type]
    except Exception:
        return default


def make_cache_fingerprint(record: dict) -> str:
    source = record.get("source")
    source_obj = source if isinstance(source, dict) else {}
    source_kind = str(source_obj.get("kind") or "")
    source_saved_at = str(source_obj.get("savedAt") or "")
    filename = str(record.get("filename") or "")
    sn_rows = record.get("snRows")
    sw_rows = record.get("swRows")
    sn_count = len(sn_rows) if isinstance(sn_rows, list) else 0
    sw_count = len(sw_rows) if isinstance(sw_rows, list) else 0
    return "|".join([source_kind, filename, source_saved_at, str(sn_count), str(sw_count)])


def normalize_cache_record(data: object) -> dict | None:
    src = data
    if not isinstance(src, dict):
        return None
    if isinstance(src.get("record"), dict):
        src = src["record"]
    if not isinstance(src, dict):
        return None

    sn_rows = src.get("snRows")
    sw_rows = src.get("swRows")
    if not isinstance(sn_rows, list):
        sn_rows = []
    if not isinstance(sw_rows, list):
        sw_rows = []
    if len(sn_rows) == 0 and len(sw_rows) == 0:
        return None
    schema_version = to_int(src.get("schemaVersion"), 0)
    if schema_version < CACHE_SCHEMA_VERSION:
        return None
    # 구버전 클라이언트 캐시(사외 유형상세 누락)는 날짜/유형 집계 오류를 유발하므로 서버 저장에서 제외한다.
    if sw_rows:
        has_legacy_sw_schema = any(
            isinstance(row, dict) and "유형상세" not in row
            for row in sw_rows
        )
        if has_legacy_sw_schema:
            return None

    source = src.get("source")
    source_obj = source if isinstance(source, dict) else {}
    source_saved_at = str(source_obj.get("savedAt") or "").strip() or utc_now_iso()
    source_kind = str(source_obj.get("kind") or "uploaded_excel_synced").strip() or "uploaded_excel_synced"
    source_label = str(source_obj.get("label") or "Shared LAN dataset").strip() or "Shared LAN dataset"

    normalized: dict = {
        "schemaVersion": CACHE_SCHEMA_VERSION,
        "version": max(0, to_int(src.get("version"), 0)),
        "updatedAt": max(0, to_int(src.get("updatedAt"), 0)),
        "filename": str(src.get("filename") or "uploaded_data"),
        "validation": src.get("validation") if isinstance(src.get("validation"), dict) else {"saenae": None, "saewae": None},
        "snRows": sn_rows,
        "swRows": sw_rows,
        "source": {
            "kind": source_kind,
            "label": source_label,
            "savedAt": source_saved_at,
        },
    }
    fp = str(src.get("fingerprint") or "").strip()
    normalized["fingerprint"] = fp or make_cache_fingerprint(normalized)
    return normalized


def load_cache_record(path: Path) -> dict | None:
    if not path.exists():
        return None
    try:
        parsed = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None
    return normalize_cache_record(parsed)


def save_cache_record(path: Path, record: dict) -> None:
    _atomic_write_json(path, record)


def _to_float(value: object, default: float = 0.0) -> float:
    try:
        if value is None:
            return default
        if isinstance(value, bool):
            return float(int(value))
        if isinstance(value, (int, float)):
            out = float(value)
            if out != out:  # NaN
                return default
            return out
        text = str(value).strip().replace(",", "").replace("%", "")
        if not text:
            return default
        out = float(text)
        if out != out:  # NaN
            return default
        return out
    except Exception:
        return default


def _build_defect_rate_export_xlsx(payload: dict) -> tuple[bytes, str]:
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

    from openpyxl.chart import LineChart, PieChart
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.marker import DataPoint
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    # 대표 보고용: 고객사 구분 색 (첫 색은 전체·축 강조용)
    _exec_palette = ["475569", "2563EB", "DC2626", "16A34A", "7C3AED", "EA580C", "0891B2"]

    def _chart_bar_value_labels(chart: BarChart) -> None:
        lbls = DataLabelList()
        lbls.showVal = True
        lbls.showLegendKey = False
        chart.dLbls = lbls

    def _chart_pie_exec_labels(chart: PieChart, *, compact: bool = False) -> None:
        lbls = DataLabelList()
        lbls.showVal = True
        lbls.showCatName = not compact
        lbls.showLegendKey = False
        chart.dLbls = lbls

    def _chart_color_bar_categories(chart: BarChart, count: int, palette_offset: int = 1) -> None:
        if not chart.series or count <= 0:
            return
        ser = chart.series[0]
        for i in range(count):
            dp = DataPoint(idx=i)
            dp.graphicalProperties.solidFill = _exec_palette[
                (i + palette_offset) % len(_exec_palette)
            ]
            ser.dPt.append(dp)

    def _chart_color_pie_slices(chart: PieChart, count: int, palette_offset: int = 1) -> None:
        if not chart.series or count <= 0:
            return
        ser = chart.series[0]
        for i in range(count):
            dp = DataPoint(idx=i)
            dp.graphicalProperties.solidFill = _exec_palette[
                (i + palette_offset) % len(_exec_palette)
            ]
            ser.dPt.append(dp)

    rows_src = payload.get("rows")
    if not isinstance(rows_src, list) or not rows_src:
        raise ValueError("rows is required.")
    rows: list[dict] = [r for r in rows_src if isinstance(r, dict)]
    if not rows:
        raise ValueError("rows is empty.")

    period_label = str(payload.get("periodLabel") or "").strip() or "전체"
    mode_label = str(payload.get("modeLabel") or "").strip() or "수량"
    source_label = str(payload.get("sourceLabel") or "").strip() or "실시간 집계"
    top_label = str(payload.get("filterLabel") or "").strip() or "거래량 상위 3개 고객"
    report_type = str(payload.get("reportType") or "executive").strip().lower() or "executive"
    max_trade = int(round(_to_float(payload.get("maxTrade"), 0.0)))
    top_customers = str(payload.get("topCustomers") or "").strip()
    original_name = str(payload.get("originalFilename") or "").strip() or "uploaded_data"
    file_name_raw = str(payload.get("fileName") or "").strip() or (
        f"신성텍_주요3사_고객불만_보고서_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )
    export_file_name = sanitize_filename(file_name_raw)
    if not export_file_name.lower().endswith(".xlsx"):
        export_file_name = f"{export_file_name}.xlsx"

    def _target_rate() -> float | None:
        raw = payload.get("targetRate")
        snap = payload.get("exportSnapshot")
        if raw is None and isinstance(snap, dict):
            raw = snap.get("targetRate")
        x = _to_float(raw, float("nan"))
        if x != x or x <= 0:
            return None
        return float(x)

    def _fmt_rate_cell(v: object) -> float | None:
        if v is None:
            return None
        x = _to_float(v, float("nan"))
        if x != x:
            return None
        return round(float(x), 3)

    target_v = _target_rate()
    snap = payload.get("exportSnapshot") if isinstance(payload.get("exportSnapshot"), dict) else None

    def _avg(values: list[float | None]) -> float | None:
        nums = [x for x in values if isinstance(x, (int, float))]
        if not nums:
            return None
        return round(sum(float(x) for x in nums) / len(nums), 3)

    def _status_by_target(rate: float | None) -> str:
        if rate is None:
            return "데이터없음"
        if target_v is None:
            return "기준없음"
        if rate > target_v:
            return "초과"
        if rate > target_v * 0.8:
            return "주의"
        return "양호"

    period_labels: list[str] = []
    total_rates: list[float | None] = []
    customer_series: list[dict] = []
    if snap:
        labels_src = snap.get("periodLabels")
        if isinstance(labels_src, list):
            period_labels = [str(x).strip() for x in labels_src if str(x).strip()]
        totals_src = snap.get("totalRates")
        if isinstance(totals_src, list):
            total_rates = [_fmt_rate_cell(x) for x in totals_src]
        series_src = snap.get("customerSeries")
        if isinstance(series_src, list):
            for item in series_src:
                if not isinstance(item, dict):
                    continue
                name = str(item.get("name") or "").strip()
                rates_src = item.get("rates")
                if not name or not isinstance(rates_src, list):
                    continue
                customer_series.append(
                    {
                        "name": name,
                        "rates": [_fmt_rate_cell(x) for x in rates_src],
                    }
                )

    if period_labels and len(total_rates) < len(period_labels):
        total_rates.extend([None] * (len(period_labels) - len(total_rates)))
    if total_rates and len(total_rates) > len(period_labels):
        total_rates = total_rates[: len(period_labels)]
    if not period_labels and total_rates:
        period_labels = [f"구간{i + 1}" for i in range(len(total_rates))]

    series_len = min(len(period_labels), len(total_rates)) if period_labels and total_rates else 0
    if series_len:
        period_labels = period_labels[:series_len]
        total_rates = total_rates[:series_len]
        for item in customer_series:
            rates = item.get("rates") if isinstance(item.get("rates"), list) else []
            if len(rates) < series_len:
                rates = rates + [None] * (series_len - len(rates))
            item["rates"] = rates[:series_len]
    else:
        period_labels = []
        total_rates = []

    def _pick_rate(values: list[float | None], idx: int | None) -> float | None:
        if idx is None or idx < 0 or idx >= len(values):
            return None
        return values[idx]

    monthly_indices: list[int] = []
    if series_len:
        monthly_indices = [i for i, label in enumerate(period_labels) if "TOTAL" not in label.upper()]
        if not monthly_indices:
            monthly_indices = list(range(series_len))

    def _has_any_rate(idx: int) -> bool:
        if _pick_rate(total_rates, idx) is not None:
            return True
        for item in customer_series:
            rates = item.get("rates") if isinstance(item.get("rates"), list) else []
            if _pick_rate(rates, idx) is not None:
                return True
        return False

    effective_monthly_indices = [idx for idx in monthly_indices if _has_any_rate(idx)]
    if not effective_monthly_indices:
        effective_monthly_indices = [idx for idx in range(series_len) if _has_any_rate(idx)]
    if not effective_monthly_indices and series_len:
        effective_monthly_indices = [0]

    current_idx = effective_monthly_indices[-1] if effective_monthly_indices else None
    prev_idx = effective_monthly_indices[-2] if len(effective_monthly_indices) >= 2 else None

    kpi_rows: list[dict] = []
    if series_len:
        overall_current = _pick_rate(total_rates, current_idx)
        overall_prev = _pick_rate(total_rates, prev_idx)
        overall_delta = (
            round(overall_current - overall_prev, 3)
            if overall_current is not None and overall_prev is not None
            else None
        )
        overall_ytd = _avg([_pick_rate(total_rates, idx) for idx in effective_monthly_indices])
        overall_gap = (
            round(overall_current - target_v, 3)
            if overall_current is not None and target_v is not None
            else None
        )
        kpi_rows.append(
            {
                "name": "전체(상위3사 합산)",
                "current": overall_current,
                "delta": overall_delta,
                "ytd": overall_ytd,
                "target": target_v,
                "gap": overall_gap,
                "status": _status_by_target(overall_current),
            }
        )
        for item in customer_series:
            rates = item.get("rates") if isinstance(item.get("rates"), list) else []
            current = _pick_rate(rates, current_idx)
            prev = _pick_rate(rates, prev_idx)
            delta = round(current - prev, 3) if current is not None and prev is not None else None
            ytd = _avg([_pick_rate(rates, idx) for idx in effective_monthly_indices])
            gap = round(current - target_v, 3) if current is not None and target_v is not None else None
            kpi_rows.append(
                {
                    "name": str(item.get("name") or ""),
                    "current": current,
                    "delta": delta,
                    "ytd": ytd,
                    "target": target_v,
                    "gap": gap,
                    "status": _status_by_target(current),
                }
            )

    comment_lines: list[str] = []
    customer_only = [x for x in kpi_rows if str(x.get("name") or "") != "전체(상위3사 합산)"]
    valid_current = [x for x in customer_only if isinstance(x.get("current"), (int, float))]
    if valid_current:
        high_row = max(valid_current, key=lambda x: float(x.get("current") or 0.0))
        if target_v is not None and isinstance(high_row.get("gap"), (int, float)):
            comment_lines.append(
                f"주의 고객사: {high_row['name']} {float(high_row['current']):.3f}% "
                f"(목표 대비 {float(high_row['gap']):+.3f}%p)"
            )
        else:
            comment_lines.append(f"주의 고객사: {high_row['name']} {float(high_row['current']):.3f}%")
    else:
        comment_lines.append("주의 고객사: 데이터 없음")

    valid_delta = [x for x in customer_only if isinstance(x.get("delta"), (int, float))]
    if valid_delta:
        up_row = max(valid_delta, key=lambda x: float(x.get("delta") or 0.0))
        down_row = min(valid_delta, key=lambda x: float(x.get("delta") or 0.0))
        comment_lines.append(f"전월 대비 상승: {up_row['name']} {abs(float(up_row['delta'])):.3f}%p")
        comment_lines.append(f"전월 대비 개선: {down_row['name']} {abs(float(down_row['delta'])):.3f}%p")
    else:
        comment_lines.append("전월 대비 상승: 비교 데이터 없음")
        comment_lines.append("전월 대비 개선: 비교 데이터 없음")

    if target_v is not None and valid_current:
        over_rows = [x for x in valid_current if float(x.get("current") or 0.0) > target_v]
        if over_rows:
            over_rows.sort(key=lambda x: float(x.get("current") or 0.0), reverse=True)
            action_targets = ", ".join(str(x.get("name") or "") for x in over_rows[:2])
            comment_lines.append(f"즉시 조치: {action_targets} 원인 Top3 점검 및 주간 개선현황 보고")
        else:
            lead = max(valid_current, key=lambda x: float(x.get("current") or 0.0))
            comment_lines.append(
                f"유지 조치: {lead['name']} 중심으로 현행 관리수준 유지, 월간 재발 모니터링 강화"
            )
    else:
        comment_lines.append("조치 포인트: 목표치/당월 데이터 확인 후 즉시개선 항목 재선정")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    thin = Side(style="thin", color="CBD5E1")
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(bold=True, size=18, color="0F172A")
    sub_font = Font(size=10, color="334155")
    section_fill = PatternFill("solid", fgColor="E2E8F0")
    over_fill = PatternFill("solid", fgColor="FEE2E2")
    warn_fill = PatternFill("solid", fgColor="FEF3C7")
    good_fill = PatternFill("solid", fgColor="DCFCE7")
    neutral_fill = PatternFill("solid", fgColor="E5E7EB")

    wb = Workbook()
    ws_cover = wb.active
    ws_cover.title = "요약보고서"
    ws_cover.merge_cells("A1:G1")
    ws_cover["A1"] = "주요 고객 3사 불량율 경영 요약"
    ws_cover["A1"].font = title_font
    ws_cover["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws_cover.row_dimensions[1].height = 30
    ws_cover.merge_cells("A2:G2")
    ws_cover["A2"] = (
        f"기준기간: {period_label} | 집계모드: {mode_label} | "
        f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    ws_cover["A2"].font = sub_font
    ws_cover["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws_cover.merge_cells("A3:G3")
    ws_cover["A3"] = (
        f"집중 고객사: {top_customers or '—'} | 목표 불량율: "
        f"{f'{target_v:.3f}%' if target_v is not None else '미설정'} | "
        f"보고서 유형: {'대표 보고용' if report_type == 'executive' else report_type}"
    )
    ws_cover["A3"].font = sub_font
    ws_cover["A3"].alignment = Alignment(horizontal="left", vertical="center")

    summary_headers = ["구분", "당월(%)", "전월대비(%p)", "YTD 평균(%)", "목표(%)", "목표대비(%p)", "상태"]
    for c, title in enumerate(summary_headers, start=1):
        cell = ws_cover.cell(row=5, column=c, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = grid_border

    data_start_row = 6
    for offset, row in enumerate(kpi_rows):
        rr = data_start_row + offset
        ws_cover.cell(row=rr, column=1, value=row.get("name") or "")
        ws_cover.cell(row=rr, column=2, value=row.get("current"))
        ws_cover.cell(row=rr, column=3, value=row.get("delta"))
        ws_cover.cell(row=rr, column=4, value=row.get("ytd"))
        ws_cover.cell(row=rr, column=5, value=row.get("target"))
        ws_cover.cell(row=rr, column=6, value=row.get("gap"))
        ws_cover.cell(row=rr, column=7, value=row.get("status") or "")
        for c in range(1, 8):
            cell = ws_cover.cell(row=rr, column=c)
            cell.border = grid_border
            cell.alignment = Alignment(horizontal="left" if c == 1 else "center", vertical="center")
            if c in (2, 3, 4, 5, 6) and isinstance(cell.value, (int, float)):
                cell.number_format = '0.000"%"'
        status = str(row.get("status") or "")
        status_fill = neutral_fill
        if status == "초과":
            status_fill = over_fill
        elif status == "주의":
            status_fill = warn_fill
        elif status == "양호":
            status_fill = good_fill
        for c in (2, 6, 7):
            ws_cover.cell(row=rr, column=c).fill = status_fill

    summary_end = data_start_row + max(len(kpi_rows) - 1, 0)
    comment_header_row = summary_end + 2
    ws_cover.merge_cells(f"A{comment_header_row}:G{comment_header_row}")
    ws_cover.cell(row=comment_header_row, column=1, value="핵심 코멘트")
    ws_cover.cell(row=comment_header_row, column=1).font = Font(bold=True, size=11, color="0F172A")
    ws_cover.cell(row=comment_header_row, column=1).fill = section_fill
    ws_cover.cell(row=comment_header_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    for i in range(4):
        rr = comment_header_row + 1 + i
        ws_cover.merge_cells(f"A{rr}:G{rr}")
        ws_cover.cell(row=rr, column=1, value=f"- {comment_lines[i] if i < len(comment_lines) else ''}")
        ws_cover.cell(row=rr, column=1).font = Font(size=10, color="1F2937")
        ws_cover.cell(row=rr, column=1).alignment = Alignment(horizontal="left", vertical="center")

    ws_cover.column_dimensions["A"].width = 24
    ws_cover.column_dimensions["B"].width = 13
    ws_cover.column_dimensions["C"].width = 14
    ws_cover.column_dimensions["D"].width = 13
    ws_cover.column_dimensions["E"].width = 11
    ws_cover.column_dimensions["F"].width = 14
    ws_cover.column_dimensions["G"].width = 10

    chart_title_row = comment_header_row + 6
    ws_cover.merge_cells(f"A{chart_title_row}:N{chart_title_row}")
    ws_cover.cell(row=chart_title_row, column=1, value="차트 요약 (막대·꺾은선·원형 · 대표 보고용)")
    ws_cover.cell(row=chart_title_row, column=1).font = Font(bold=True, size=11, color="0F172A")
    ws_cover.cell(row=chart_title_row, column=1).fill = section_fill
    ws_cover.cell(row=chart_title_row, column=1).alignment = Alignment(horizontal="left", vertical="center")

    chart_col = 30  # AD (인쇄영역 바깥)
    chart_headers = ["기간", "전체(상위3사 합산)"] + [str(x.get("name") or "") for x in customer_series]
    if target_v is not None:
        chart_headers.append("목표부적합률(%)")
    for c, title in enumerate(chart_headers, start=chart_col):
        cell = ws_cover.cell(row=2, column=c, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = grid_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    chart_indices: list[int] = []
    if series_len:
        chart_indices = [idx for idx in range(series_len) if _has_any_rate(idx)]
        if not chart_indices:
            chart_indices = list(range(series_len))
        if len(chart_indices) > 12:
            chart_indices = chart_indices[-12:]
    for out_row, src_idx in enumerate(chart_indices, start=3):
        values: list[object] = [period_labels[src_idx], total_rates[src_idx]]
        for item in customer_series:
            rates = item.get("rates") if isinstance(item.get("rates"), list) else []
            values.append(rates[src_idx] if src_idx < len(rates) else None)
        if target_v is not None:
            values.append(round(target_v, 3))
        for c_off, val in enumerate(values):
            cell = ws_cover.cell(row=out_row, column=chart_col + c_off, value=val)
            cell.border = grid_border
            if c_off > 0 and isinstance(val, (int, float)):
                cell.number_format = '0.000"%"'
            cell.alignment = Alignment(horizontal="center", vertical="center")

    customer_current_rows = [x for x in customer_only if isinstance(x.get("current"), (int, float))]
    customer_current_rows.sort(key=lambda x: float(x.get("current") or 0.0), reverse=True)
    bad_cnt_map: dict[str, int] = {}
    for row in rows:
        name = str(row.get("고객사") or "").strip()
        bad_cnt = int(round(_to_float(row.get("부적합건수"), 0.0)))
        if name:
            bad_cnt_map[name] = bad_cnt
    total_bad_cnt = sum(bad_cnt_map.values())
    for item in customer_current_rows:
        name = str(item.get("name") or "")
        bad_cnt = int(bad_cnt_map.get(name, 0))
        share = (float(bad_cnt) / total_bad_cnt * 100.0) if total_bad_cnt > 0 else 0.0
        item["bad_cnt"] = bad_cnt
        item["share"] = round(share, 3)
        cur = item.get("current")
        delta = item.get("delta")
        if isinstance(cur, (int, float)) and isinstance(delta, (int, float)):
            item["prev"] = round(float(cur) - float(delta), 3)
        else:
            item["prev"] = None

    cust_chart_col = chart_col + len(chart_headers) + 2
    cust_headers = ["고객사", "당월부적합률(%)", "목표부적합률(%)", "전월부적합률(%)", "전월대비(%p)", "기여율(%)"]
    for c, title in enumerate(cust_headers, start=cust_chart_col):
        cell = ws_cover.cell(row=2, column=c, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = grid_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for idx, item in enumerate(customer_current_rows, start=3):
        ws_cover.cell(row=idx, column=cust_chart_col, value=str(item.get("name") or ""))
        ws_cover.cell(row=idx, column=cust_chart_col + 1, value=float(item.get("current") or 0.0))
        ws_cover.cell(row=idx, column=cust_chart_col + 2, value=round(target_v, 3) if target_v is not None else None)
        ws_cover.cell(row=idx, column=cust_chart_col + 3, value=item.get("prev"))
        ws_cover.cell(row=idx, column=cust_chart_col + 4, value=item.get("delta"))
        ws_cover.cell(row=idx, column=cust_chart_col + 5, value=item.get("share"))
        for c_idx in range(cust_chart_col, cust_chart_col + len(cust_headers)):
            cell = ws_cover.cell(row=idx, column=c_idx)
            cell.border = grid_border
            if c_idx > cust_chart_col and isinstance(cell.value, (int, float)):
                cell.number_format = '0.000"%"'

    chart_top_row = chart_title_row + 1
    # 한 페이지 세로 맞춤(fitToHeight)은 차트를 세로로 찌그러뜨림 → 꺾은선·막대·원형은 아래 줄에 나란히 배치
    charts_row_line = chart_top_row
    charts_row_bar_pie = chart_top_row + 34

    if chart_indices:
        line_cover = LineChart()
        line_cover.title = f"월별 부적합율 추이 (%) · {period_label}"
        line_cover.style = 13
        line_cover.y_axis.title = "부적합율 (%)"
        line_cover.x_axis.title = "기간"
        line_cover.height = 9.0
        line_cover.width = 24.0
        line_cover.legend.position = "r"
        line_cover.legend.overlay = False
        line_cover.plot_visible_only = False
        line_cover.y_axis.majorGridlines = None
        line_cover.y_axis.scaling.min = 0
        rate_pool = [v for v in total_rates if isinstance(v, (int, float))]
        for item in customer_series:
            rates = item.get("rates") if isinstance(item.get("rates"), list) else []
            rate_pool.extend([v for v in rates if isinstance(v, (int, float))])
        if target_v is not None:
            rate_pool.append(target_v)
        y_max = max(rate_pool) if rate_pool else 1.0
        line_cover.y_axis.scaling.max = max(1.0, round(float(y_max) * 1.25, 3))
        data_ref = Reference(
            ws_cover,
            min_col=chart_col + 1,
            min_row=2,
            max_col=chart_col + len(chart_headers) - 1,
            max_row=2 + len(chart_indices),
        )
        cats_ref = Reference(ws_cover, min_col=chart_col, min_row=3, max_row=2 + len(chart_indices))
        line_cover.add_data(data_ref, titles_from_data=True)
        line_cover.set_categories(cats_ref)
        line_palette = ["334155", "2563EB", "DC2626", "16A34A", "7C3AED", "0EA5E9"]
        for s_idx, ser in enumerate(line_cover.series):
            color = line_palette[s_idx % len(line_palette)]
            if target_v is not None and s_idx == len(line_cover.series) - 1:
                color = "F59E0B"
                ser.graphicalProperties.line.dashStyle = "sysDot"
                ser.marker.symbol = "none"
            else:
                ser.marker.symbol = "circle"
                ser.marker.size = 7
            ser.graphicalProperties.line.solidFill = color
            ser.graphicalProperties.line.width = 28500
            ser.smooth = True
        ws_cover.add_chart(line_cover, f"A{charts_row_line}")

    if customer_current_rows:
        bar = BarChart()
        bar.type = "col"
        bar.style = 13
        bar.title = "당월 고객사 vs 목표 부적합율 (%)"
        bar.y_axis.title = "부적합율 (%)"
        bar.x_axis.title = "고객사"
        bar.height = 8.0
        bar.width = 14.0
        bar.gapWidth = 48
        bar.plot_visible_only = False
        bar.y_axis.scaling.min = 0
        max_current = max(float(x.get("current") or 0.0) for x in customer_current_rows)
        ref_max = max(max_current, target_v if target_v is not None else 0.0)
        bar.y_axis.scaling.max = max(1.0, round(ref_max * 1.28, 3))
        bar_data_max = 2 + len(customer_current_rows)
        bar_cats = Reference(ws_cover, min_col=cust_chart_col, min_row=3, max_row=bar_data_max)
        bar.add_data(
            Reference(
                ws_cover,
                min_col=cust_chart_col + 1,
                min_row=2,
                max_col=cust_chart_col + 1,
                max_row=bar_data_max,
            ),
            titles_from_data=True,
        )
        if target_v is not None:
            bar.add_data(
                Reference(
                    ws_cover,
                    min_col=cust_chart_col + 2,
                    min_row=2,
                    max_col=cust_chart_col + 2,
                    max_row=bar_data_max,
                ),
                titles_from_data=True,
            )
        bar.set_categories(bar_cats)
        _chart_color_bar_categories(bar, len(customer_current_rows))
        if target_v is not None and len(bar.series) >= 2:
            bar.series[1].graphicalProperties.solidFill = "FBBF24"
            bar.series[1].graphicalProperties.line.solidFill = "D97706"
        _chart_bar_value_labels(bar)
        ws_cover.add_chart(bar, f"A{charts_row_bar_pie}")

    if customer_current_rows:
        pie = PieChart()
        pie.title = "고객사별 부적합건수 기여율"
        pie.style = 13
        pie.height = 8.0
        pie.width = 12.0
        pie.legend.position = "r"
        pie.legend.overlay = False
        pie.plot_visible_only = False
        pie_max_row = 2 + len(customer_current_rows)
        pie_data = Reference(
            ws_cover,
            min_col=cust_chart_col + 5,  # 기여율(%)
            min_row=2,
            max_col=cust_chart_col + 5,
            max_row=pie_max_row,
        )
        pie_cats = Reference(
            ws_cover,
            min_col=cust_chart_col,  # 고객사
            min_row=3,
            max_row=pie_max_row,
        )
        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(pie_cats)
        _chart_color_pie_slices(pie, len(customer_current_rows))
        _chart_pie_exec_labels(pie, compact=True)
        ws_cover.add_chart(pie, f"I{charts_row_bar_pie}")

    priority_row = charts_row_bar_pie + 28
    ws_cover.merge_cells(f"A{priority_row}:N{priority_row}")
    ws_cover.cell(row=priority_row, column=1, value="우선순위 요약")
    ws_cover.cell(row=priority_row, column=1).font = Font(bold=True, size=11, color="0F172A")
    ws_cover.cell(row=priority_row, column=1).fill = section_fill
    ws_cover.cell(row=priority_row, column=1).alignment = Alignment(horizontal="left", vertical="center")

    priority_lines: list[str] = []
    if customer_current_rows:
        risk = customer_current_rows[0]
        gap = risk.get("gap")
        gap_text = f"{float(gap):+.3f}%p" if isinstance(gap, (int, float)) else "비교불가"
        priority_lines.append(
            f"최고 위험: {risk['name']} 당월 {float(risk['current']):.3f}% (목표 대비 {gap_text})"
        )
        top_share = max(customer_current_rows, key=lambda x: float(x.get("share") or 0.0))
        priority_lines.append(
            f"기여도 1위: {top_share['name']} {float(top_share.get('share') or 0.0):.3f}%"
        )
    delta_numeric = [x for x in customer_current_rows if isinstance(x.get("delta"), (int, float))]
    if delta_numeric:
        worst = max(delta_numeric, key=lambda x: float(x.get("delta") or 0.0))
        best = min(delta_numeric, key=lambda x: float(x.get("delta") or 0.0))
        if float(worst.get("delta") or 0.0) > 0:
            priority_lines.append(
                f"전월 악화: {worst['name']} +{float(worst['delta']):.3f}%p (원인 재점검 필요)"
            )
        else:
            priority_lines.append("전월 악화: 없음")
        if float(best.get("delta") or 0.0) < 0:
            priority_lines.append(
                f"최대 개선: {best['name']} {float(best['delta']):.3f}%p (개선활동 유지)"
            )
        else:
            priority_lines.append("최대 개선: 없음")
    if not priority_lines:
        priority_lines.append("우선순위 데이터 없음")

    for i, text in enumerate(priority_lines[:4]):
        rr = priority_row + 1 + i
        ws_cover.merge_cells(f"A{rr}:N{rr}")
        ws_cover.cell(row=rr, column=1, value=f"- {text}")
        ws_cover.cell(row=rr, column=1).font = Font(size=10, color="1F2937")
        ws_cover.cell(row=rr, column=1).alignment = Alignment(horizontal="left", vertical="center")

    ws_cover.page_setup.orientation = ws_cover.ORIENTATION_LANDSCAPE
    ws_cover.page_setup.paperSize = ws_cover.PAPERSIZE_A4
    ws_cover.page_setup.fitToWidth = 1
    ws_cover.page_setup.fitToHeight = 0
    ws_cover.page_margins = PageMargins(left=0.25, right=0.25, top=0.4, bottom=0.4, header=0.2, footer=0.2)
    ws_cover.print_area = f"A1:N{max(priority_row + 28, 72)}"
    ws_cover.oddHeader.left.text = "&\"맑은 고딕,Bold\"&11 신성텍 품질경영 보고서"
    ws_cover.oddHeader.right.text = "&D &T"
    ws_cover.oddFooter.center.text = "&P / &N"

    ws_m = wb.create_sheet("월별추이")
    if series_len:
        monthly_chart_indices = [idx for idx in range(series_len) if _has_any_rate(idx)]
        if not monthly_chart_indices:
            monthly_chart_indices = list(range(series_len))
        heads = ["기간", "전체(상위3사 합산)"] + [str(x.get("name") or "") for x in customer_series]
        if target_v is not None:
            heads.append("목표부적합률(%)")
        ws_m.append(heads)
        for cell in ws_m[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = grid_border
        for src_idx in monthly_chart_indices:
            row_vals: list[object] = [period_labels[src_idx], _fmt_rate_cell(total_rates[src_idx])]
            for item in customer_series:
                rates = item.get("rates") if isinstance(item.get("rates"), list) else []
                row_vals.append(_fmt_rate_cell(rates[src_idx] if src_idx < len(rates) else None))
            if target_v is not None:
                row_vals.append(round(target_v, 3))
            ws_m.append(row_vals)
        last_data_row = ws_m.max_row
        for r in ws_m.iter_rows(min_row=2, max_row=last_data_row, min_col=1, max_col=len(heads)):
            for cell in r:
                cell.border = grid_border
                if cell.column > 1 and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.000"%"'
        for col_idx in range(1, len(heads) + 1):
            ws_m.column_dimensions[get_column_letter(col_idx)].width = 14 if col_idx == 1 else 12
        ws_m.freeze_panes = "B2"

        appendix_start = last_data_row + 2
        ws_m.cell(row=appendix_start, column=1, value="(참고) 기간별 출하·부적합 건수")
        ws_m.cell(row=appendix_start, column=1).font = Font(bold=True, size=11, color="0F172A")
        pm = snap.get("periodMetrics") if isinstance(snap.get("periodMetrics") if snap else None, list) else []
        h2: list[object] = ["구분"] + period_labels
        ws_m.append(h2)
        ship_row: list[object] = ["출하건수"]
        bad_row: list[object] = ["부적합건수"]
        for i in range(series_len):
            if i < len(pm) and isinstance(pm[i], dict):
                ship_row.append(int(round(_to_float(pm[i].get("입고건수"), 0.0))))
                bad_row.append(int(round(_to_float(pm[i].get("부적합건수"), 0.0))))
            else:
                ship_row.append("")
                bad_row.append("")
        ws_m.append(ship_row)
        ws_m.append(bad_row)

        line = LineChart()
        line.title = f"월별 부적합율 추이 (%) · {period_label} · {mode_label} 기준"
        line.style = 13
        line.y_axis.title = "부적합율 (%)"
        line.x_axis.title = "기간"
        line.height = 11.0
        line.width = 21.0
        line.legend.position = "b"
        line.y_axis.majorGridlines = None
        line.y_axis.scaling.min = 0
        rate_pool = [v for v in total_rates if isinstance(v, (int, float))]
        for item in customer_series:
            rates = item.get("rates") if isinstance(item.get("rates"), list) else []
            rate_pool.extend([v for v in rates if isinstance(v, (int, float))])
        if target_v is not None:
            rate_pool.append(target_v)
        y_max = max(rate_pool) if rate_pool else 1.0
        line.y_axis.scaling.max = max(1.0, round(float(y_max) * 1.25, 3))
        data_ref = Reference(ws_m, min_col=2, min_row=1, max_col=len(heads), max_row=last_data_row)
        cats_ref = Reference(ws_m, min_col=1, min_row=2, max_row=last_data_row)
        line.add_data(data_ref, titles_from_data=True)
        line.set_categories(cats_ref)
        line_palette = ["334155", "2563EB", "DC2626", "16A34A", "7C3AED", "0EA5E9"]
        for s_idx, ser in enumerate(line.series):
            color = line_palette[s_idx % len(line_palette)]
            if target_v is not None and s_idx == len(line.series) - 1:
                color = "F59E0B"
                ser.graphicalProperties.line.dashStyle = "sysDot"
                ser.marker.symbol = "none"
            else:
                ser.marker.symbol = "circle"
                ser.marker.size = 7
            ser.graphicalProperties.line.solidFill = color
            ser.graphicalProperties.line.width = 28500
            ser.smooth = True
        chart_row = ws_m.max_row + 2
        ws_m.add_chart(line, f"A{chart_row}")

        bar_m = BarChart()
        bar_m.type = "col"
        bar_m.style = 13
        bar_m.title = "기간별 전체(상위3사) 부적합율 (막대) · 목표 대비"
        bar_m.y_axis.title = "부적합율 (%)"
        bar_m.x_axis.title = "기간"
        bar_m.height = 8.5
        bar_m.width = 20.5
        bar_m.gapWidth = 42
        bar_m.plot_visible_only = False
        bar_m.y_axis.scaling.min = 0
        rate_only = [v for v in total_rates if isinstance(v, (int, float))]
        bar_pool = list(rate_only)
        if target_v is not None:
            bar_pool.append(target_v)
        y_max_b = max(bar_pool) if bar_pool else 1.0
        bar_m.y_axis.scaling.max = max(1.0, round(float(y_max_b) * 1.22, 3))
        bar_m.add_data(
            Reference(ws_m, min_col=2, min_row=1, max_col=2, max_row=last_data_row),
            titles_from_data=True,
        )
        if target_v is not None:
            tgt_col = len(heads)
            bar_m.add_data(
                Reference(ws_m, min_col=tgt_col, min_row=1, max_col=tgt_col, max_row=last_data_row),
                titles_from_data=True,
            )
            if len(bar_m.series) >= 2:
                bar_m.series[1].graphicalProperties.solidFill = "FCD34D"
                bar_m.series[1].graphicalProperties.line.solidFill = "D97706"
        cats_m = Reference(ws_m, min_col=1, min_row=2, max_row=last_data_row)
        bar_m.set_categories(cats_m)
        _chart_bar_value_labels(bar_m)
        if bar_m.series:
            _chart_color_bar_categories(bar_m, last_data_row - 1, palette_offset=0)
        ws_m.add_chart(bar_m, f"A{chart_row + 28}")
    else:
        ws_m["A1"] = (
            "월별 추이 데이터가 없습니다. 대시보드 「주요3사 고객불만 현황」 탭을 연 상태에서 "
            "데이터를 불러온 뒤 다시보내기 해 주세요."
        )
        ws_m["A1"].alignment = Alignment(wrap_text=True)
    ws_m.page_setup.orientation = ws_m.ORIENTATION_LANDSCAPE
    ws_m.page_setup.paperSize = ws_m.PAPERSIZE_A4
    ws_m.page_setup.fitToWidth = 1
    ws_m.page_setup.fitToHeight = 0
    ws_m.page_margins = PageMargins(left=0.25, right=0.25, top=0.4, bottom=0.4, header=0.2, footer=0.2)
    ws_m.oddFooter.center.text = "&P / &N"

    ws = wb.create_sheet("집계")
    headers = [
        "고객사",
        "납품수량",
        "검사수량",
        "납품건수",
        "부적합수량",
        "부적합건수",
        "불량율_수량(%)",
        "불량율_건수(%)",
        "표시불량율(%)",
        "목표부적합률참고(%)",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = grid_border

    for row in rows:
        g = round(_to_float(row.get("불량율_수량(%)"), 0.0), 3)
        h = round(_to_float(row.get("불량율_건수(%)"), 0.0), 3)
        i_val = round(_to_float(row.get("표시불량율(%)"), 0.0), 3)
        tgt_cell = round(target_v, 3) if target_v is not None else None
        ws.append(
            [
                str(row.get("고객사") or ""),
                int(round(_to_float(row.get("납품수량", row.get("입고수량")), 0.0))),
                int(round(_to_float(row.get("검사수량"), 0.0))),
                int(round(_to_float(row.get("납품건수", row.get("입고건수")), 0.0))),
                int(round(_to_float(row.get("부적합수량"), 0.0))),
                int(round(_to_float(row.get("부적합건수"), 0.0))),
                g,
                h,
                i_val,
                tgt_cell,
            ]
        )

    ws.column_dimensions["A"].width = 18
    for col in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[col].width = 12
    for col in ("G", "H", "I", "J"):
        ws.column_dimensions[col].width = 14

    row_end = ws.max_row
    for r_idx in range(2, row_end + 1):
        for c_idx in range(1, 11):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = grid_border
            if c_idx in (7, 8, 9, 10) and isinstance(cell.value, (int, float)):
                cell.number_format = '0.000"%"'
        if target_v is not None:
            rate_v = ws.cell(row=r_idx, column=9).value
            if isinstance(rate_v, (int, float)):
                status_fill = good_fill
                if float(rate_v) > target_v:
                    status_fill = over_fill
                elif float(rate_v) > target_v * 0.8:
                    status_fill = warn_fill
                ws.cell(row=r_idx, column=9).fill = status_fill

    if row_end >= 2:
        bc = BarChart()
        bc.type = "col"
        bc.style = 13
        bc.title = "고객사별 표시 부적합율 (막대)" + (
            f" · 목표 {target_v:.3f} %" if target_v is not None else ""
        )
        bc.y_axis.title = "부적합율 (%)"
        bc.x_axis.title = "고객사"
        bc.gapWidth = 72
        bc.add_data(Reference(ws, min_col=9, min_row=1, max_col=9, max_row=row_end), titles_from_data=True)
        bc.set_categories(Reference(ws, min_col=1, min_row=2, max_row=row_end))
        bc.height = 10.0
        bc.width = 17
        _chart_color_bar_categories(bc, row_end - 1)
        _chart_bar_value_labels(bc)
        ws.add_chart(bc, "L2")

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.4, bottom=0.4, header=0.2, footer=0.2)
    ws.print_area = f"A1:J{max(row_end, 2)}"
    ws.oddFooter.center.text = "&P / &N"

    ws_raw = wb.create_sheet("원본데이터")
    ws_raw.append(headers[:-1])
    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = grid_border
    for row in rows:
        ws_raw.append(
            [
                str(row.get("고객사") or ""),
                int(round(_to_float(row.get("납품수량", row.get("입고수량")), 0.0))),
                int(round(_to_float(row.get("검사수량"), 0.0))),
                int(round(_to_float(row.get("납품건수", row.get("입고건수")), 0.0))),
                int(round(_to_float(row.get("부적합수량"), 0.0))),
                int(round(_to_float(row.get("부적합건수"), 0.0))),
                round(_to_float(row.get("불량율_수량(%)"), 0.0), 3),
                round(_to_float(row.get("불량율_건수(%)"), 0.0), 3),
                round(_to_float(row.get("표시불량율(%)"), 0.0), 3),
            ]
        )
    raw_end = ws_raw.max_row
    for r in ws_raw.iter_rows(min_row=2, max_row=raw_end, min_col=1, max_col=9):
        for cell in r:
            cell.border = grid_border
            if cell.column in (7, 8, 9) and isinstance(cell.value, (int, float)):
                cell.number_format = '0.000"%"'
    ws_raw.column_dimensions["A"].width = 18
    for col in ("B", "C", "D", "E", "F"):
        ws_raw.column_dimensions[col].width = 12
    for col in ("G", "H", "I"):
        ws_raw.column_dimensions[col].width = 14
    ws_raw.page_setup.orientation = ws_raw.ORIENTATION_LANDSCAPE
    ws_raw.page_setup.paperSize = ws_raw.PAPERSIZE_A4
    ws_raw.page_setup.fitToWidth = 1
    ws_raw.page_setup.fitToHeight = 0
    ws_raw.page_margins = PageMargins(left=0.25, right=0.25, top=0.4, bottom=0.4, header=0.2, footer=0.2)
    ws_raw.oddFooter.center.text = "&P / &N"

    ws_meta = wb.create_sheet("메타")
    ws_meta.append(["문서유형", "주요 3사 고객불만 · 경영보고용"])
    ws_meta.append(["생성일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_meta.append(["기준기간", period_label])
    ws_meta.append(["집계모드", mode_label])
    ws_meta.append(["보고서유형", report_type])
    ws_meta.append(["목표 부적합율(%)", f"{target_v:.3f}" if target_v is not None else "미설정"])
    ws_meta.append(["데이터소스", source_label])
    ws_meta.append(["내보내기 필터", top_label])
    ws_meta.append(["최대 거래량", max_trade])
    ws_meta.append(["보낸 고객 수", len(rows)])
    if top_customers:
        ws_meta.append(["보낸 고객사", top_customers])
    ws_meta.append(["원본파일명", original_name])
    if snap:
        y = snap.get("year")
        if y:
            ws_meta.append(["집계 연도(스냅샷)", str(y)])
        tab = snap.get("customerTab")
        if tab:
            ws_meta.append(["대시보드 고객 탭", str(tab)])
    for r in ws_meta.iter_rows(min_row=1, max_row=ws_meta.max_row, min_col=1, max_col=2):
        for cell in r:
            cell.border = grid_border
    ws_meta.column_dimensions["A"].width = 22
    ws_meta.column_dimensions["B"].width = 48

    buff = BytesIO()
    wb.save(buff)
    return buff.getvalue(), export_file_name


def _token_equal(a: str, b: str) -> bool:
    """타이밍 안전 비교. 길이가 다르면 False."""
    left = str(a or "").encode("utf-8")
    right = str(b or "").encode("utf-8")
    if len(left) != len(right):
        return False
    return hmac.compare_digest(left, right)


def read_public_tunnel_url(path: Path) -> str:
    if not path.exists():
        return ""
    try:
        text = path.read_text(encoding="utf-8").strip()
    except OSError:
        return ""
    # 첫 줄만 사용 (주석/부가 정보 무시)
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("http://") or line.startswith("https://"):
            return line.rstrip("/")
    return ""


class SharedAttachmentServer(ThreadingHTTPServer):
    daemon_threads = True
    # Windows에서 중복 프로세스가 같은 포트를 점유하는 현상을 방지한다.
    allow_reuse_address = False

    def __init__(
        self,
        server_address,
        request_handler_class,
        repo_root: Path,
        allow_public_clients: bool = False,
        access_token: str = "",
    ):
        super().__init__(server_address, request_handler_class)
        self.repo_root = repo_root
        self.cloud_dir = repo_root / "cloud"
        self.attach_dir = self.cloud_dir / "sw_attachments"
        self.manifest_path = self.cloud_dir / "sw_attachments_manifest.json"
        self.cache_path = self.cloud_dir / "uploaded_report_cache.json"
        self.public_url_path = self.cloud_dir / "public_tunnel_url.txt"
        # 멀티 스레드 요청에서 캐시/manifest 동시 쓰기 충돌 방지
        self.cache_lock = threading.Lock()
        self.manifest_lock = threading.Lock()
        # gzip 압축 결과 캐시 (파일 경로+mtime+size 기준)
        self.gzip_cache_lock = threading.Lock()
        self.gzip_cache: OrderedDict[str, bytes] = OrderedDict()
        self.allow_public_clients = bool(allow_public_clients)
        self.access_token = str(access_token or "").strip()
        self.token_required = bool(self.access_token)


class SharedAttachmentHandler(SimpleHTTPRequestHandler):
    def _client_ip(self) -> str:
        host = self.client_address[0] if self.client_address else ""
        return str(host or "")

    def _is_private_or_local_ip(self, host: str) -> bool:
        # 기본 운영 정책: 내부망/로컬 대역만 허용
        try:
            ip_obj = ipaddress.ip_address(host)
            if isinstance(ip_obj, ipaddress.IPv6Address) and ip_obj.ipv4_mapped:
                ip_obj = ip_obj.ipv4_mapped
        except ValueError:
            return False
        return bool(
            ip_obj.is_private
            or ip_obj.is_loopback
            or ip_obj.is_link_local
            or ip_obj.is_reserved
        )

    def _is_client_allowed(self) -> bool:
        # 운영자가 공개 허용 옵션을 준 경우는 예외적으로 전체 허용
        if getattr(self.server, "allow_public_clients", False):
            return True
        return self._is_private_or_local_ip(self._client_ip())

    def _required_access_token(self) -> str:
        return str(getattr(self.server, "access_token", "") or "").strip()

    def _extract_client_token(self) -> str:
        header = (
            self.headers.get("X-NCP-Token")
            or self.headers.get("X-Access-Token")
            or ""
        ).strip()
        if header:
            return header
        parsed = urlparse(self.path)
        qs = parse_qs(parsed.query)
        q_vals = qs.get("token") or qs.get("access_token") or []
        if q_vals and str(q_vals[0]).strip():
            return str(q_vals[0]).strip()
        cookie_raw = self.headers.get("Cookie") or ""
        for part in cookie_raw.split(";"):
            name, _, value = part.strip().partition("=")
            if name in ("ncp_token", "ncp_access_token") and value.strip():
                return value.strip()
        return ""

    def _is_token_authorized(self) -> bool:
        required = self._required_access_token()
        if not required:
            return True
        return _token_equal(self._extract_client_token(), required)

    def _wants_https_cookie(self) -> bool:
        proto = (self.headers.get("X-Forwarded-Proto") or "").strip().lower()
        if proto == "https":
            return True
        host = (self.headers.get("Host") or "").lower()
        return host.endswith("trycloudflare.com") or host.endswith("cfargotunnel.com")

    def _set_access_cookie_header(self) -> None:
        token = self._required_access_token()
        if not token:
            return
        flags = "Path=/; Max-Age=2592000; SameSite=Lax; HttpOnly"
        if self._wants_https_cookie():
            flags += "; Secure"
        self.send_header("Set-Cookie", f"ncp_token={quote(token, safe='')}; {flags}")

    def _send_token_gate_html(self) -> None:
        body = (
            "<!DOCTYPE html><html lang='ko'><head><meta charset='UTF-8'>"
            "<meta name='viewport' content='width=device-width, initial-scale=1'>"
            "<title>접속 인증 · 신성텍 부적합 보고서</title>"
            "<style>"
            "body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Malgun Gothic',sans-serif;"
            "background:#F3F4F6;margin:0;min-height:100vh;display:flex;align-items:center;justify-content:center;padding:24px;}"
            ".card{background:#fff;border:1px solid #E5E7EB;border-radius:16px;padding:28px 24px;max-width:420px;width:100%;"
            "box-shadow:0 10px 30px rgba(15,23,42,.08);}"
            "h1{font-size:18px;margin:0 0 8px;color:#1F2937;}p{font-size:13px;color:#6B7280;line-height:1.55;margin:0 0 18px;}"
            "input{width:100%;box-sizing:border-box;padding:12px 14px;border:1px solid #D1D5DB;border-radius:10px;font-size:15px;}"
            "button{margin-top:12px;width:100%;padding:12px;border:0;border-radius:10px;background:#2563EB;color:#fff;"
            "font-weight:700;font-size:14px;cursor:pointer;}"
            ".hint{margin-top:14px;font-size:11.5px;color:#9CA3AF;}"
            "</style></head><body><div class='card'>"
            "<h1>🔒 모바일/외부 접속 인증</h1>"
            "<p>서버에서 발급한 접속 토큰을 입력하세요. 사내 PC에서 "
            "<b>모바일접속_언제든지.bat</b> 실행 시 콘솔에 표시됩니다.</p>"
            "<form method='GET' action='/index.html'>"
            "<input type='password' name='token' placeholder='접속 토큰' autocomplete='current-password' required>"
            "<button type='submit'>보고서 열기</button></form>"
            "<div class='hint'>토큰은 브라우저에 30일간 저장됩니다.</div>"
            "</div></body></html>"
        ).encode("utf-8")
        self.send_response(401)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _reject_unauthorized(self, route: str) -> None:
        if route.startswith("/api/"):
            self._send_json(401, {"ok": False, "error": "Access token required."})
            return
        self._send_token_gate_html()

    def _reject_forbidden(self, route: str) -> None:
        if route.startswith("/api/"):
            self._send_json(403, {"ok": False, "error": "Client IP is not allowed."})
            return
        self.send_error(403, "Client IP is not allowed.")

    def _send_json(self, status: int, payload: dict) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        if status < 400 and self._required_access_token() and self._is_token_authorized():
            self._set_access_cookie_header()
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_binary(
        self,
        status: int,
        payload: bytes,
        content_type: str = "application/octet-stream",
        download_name: str | None = None,
    ) -> None:
        body = payload if isinstance(payload, (bytes, bytearray)) else bytes(payload)
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(body)))
        if download_name:
            safe_name = sanitize_filename(download_name)
            self.send_header(
                "Content-Disposition",
                f"attachment; filename*=UTF-8''{quote(safe_name)}",
            )
        self.end_headers()
        self.wfile.write(body)

    def _read_json_body(self) -> dict:
        length_header = self.headers.get("Content-Length", "")
        try:
            length = int(length_header)
        except ValueError:
            raise ValueError("Invalid Content-Length header.")
        if length <= 0:
            raise ValueError("Request body is empty.")
        if length > 60 * 1024 * 1024:
            raise ValueError("Request body is too large.")
        raw = self.rfile.read(length)
        try:
            parsed = json.loads(raw.decode("utf-8"))
        except Exception:
            raise ValueError("Invalid JSON body.")
        if not isinstance(parsed, dict):
            raise ValueError("JSON body must be an object.")
        return parsed

    def _client_accepts_gzip(self) -> bool:
        enc = str(self.headers.get("Accept-Encoding", "")).lower()
        return "gzip" in enc

    def _resolve_static_file_path(self, route: str) -> Path | None:
        candidate = Path(self.translate_path(route))
        if candidate.is_dir():
            for index_name in ("index.html", "index.htm"):
                index_path = candidate / index_name
                if index_path.is_file():
                    return index_path
            return None
        if candidate.is_file():
            return candidate
        return None

    def _is_gzip_candidate(self, file_path: Path, content_type: str) -> bool:
        ext = file_path.suffix.lower()
        if ext in GZIP_STATIC_EXTENSIONS:
            return True
        ctype = str(content_type or "").split(";", 1)[0].strip().lower()
        if ctype.startswith("text/"):
            return True
        return ctype in GZIP_CONTENT_TYPES

    def _get_gzip_payload(self, file_path: Path, stat_obj: os.stat_result) -> bytes | None:
        key = f"{file_path}|{int(stat_obj.st_mtime_ns)}|{int(stat_obj.st_size)}"
        cache = self.server.gzip_cache
        lock = self.server.gzip_cache_lock
        with lock:
            cached = cache.get(key)
            if cached is not None:
                cache.move_to_end(key)
                return cached
        try:
            raw = file_path.read_bytes()
        except OSError:
            return None
        if len(raw) < GZIP_MIN_BYTES:
            return None
        payload = gzip.compress(raw, compresslevel=GZIP_LEVEL)
        # 압축 이득이 거의 없으면 원본 전송 유지
        if len(payload) >= len(raw) - 64:
            return None
        with lock:
            cache[key] = payload
            cache.move_to_end(key)
            while len(cache) > GZIP_CACHE_MAX_ITEMS:
                cache.popitem(last=False)
        return payload

    def _maybe_send_gzip_static(self, route: str, send_body: bool) -> bool:
        if route.startswith("/api/"):
            return False
        if not self._client_accepts_gzip():
            return False
        if self.headers.get("Range"):
            # Range 응답은 기본 구현에 위임
            return False
        file_path = self._resolve_static_file_path(route)
        if not file_path:
            return False
        content_type = self.guess_type(str(file_path))
        if not self._is_gzip_candidate(file_path, content_type):
            return False
        try:
            stat_obj = file_path.stat()
        except OSError:
            return False
        payload = self._get_gzip_payload(file_path, stat_obj)
        if payload is None:
            return False
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Encoding", "gzip")
        self.send_header("Vary", "Accept-Encoding")
        self.send_header("Content-Length", str(len(payload)))
        self.send_header("Last-Modified", self.date_time_string(stat_obj.st_mtime))
        ext = file_path.suffix.lower()
        if ext in (".html", ".htm"):
            self.send_header("Cache-Control", "no-store, must-revalidate")
        else:
            self.send_header("Cache-Control", "no-cache")
        self.end_headers()
        if send_body:
            self.wfile.write(payload)
        return True

    def do_GET(self) -> None:
        route = urlparse(self.path).path
        # 요청 시작 단계에서 클라이언트 접근 정책을 우선 검사
        if not self._is_client_allowed():
            self._reject_forbidden(route)
            return
        if not self._is_token_authorized():
            self._reject_unauthorized(route)
            return
        if self._maybe_consume_token_query():
            return
        if route == "/api/server-info":
            self._handle_server_info()
            return
        if route == "/api/sw-attachments/manifest":
            self._handle_manifest_get()
            return
        if route == "/api/data/cache":
            self._handle_cache_get()
            return
        if self._maybe_send_gzip_static(route, send_body=True):
            return
        super().do_GET()

    def do_HEAD(self) -> None:
        route = urlparse(self.path).path
        if not self._is_client_allowed():
            self._reject_forbidden(route)
            return
        if not self._is_token_authorized():
            self._reject_unauthorized(route)
            return
        if self._maybe_send_gzip_static(route, send_body=False):
            return
        super().do_HEAD()

    def do_POST(self) -> None:
        route = urlparse(self.path).path
        # 쓰기 API 역시 동일한 접근 정책 적용
        if not self._is_client_allowed():
            self._reject_forbidden(route)
            return
        if not self._is_token_authorized():
            self._reject_unauthorized(route)
            return
        if route == "/api/sw-attachments/upload":
            self._handle_upload()
            return
        if route == "/api/sw-attachments/delete":
            self._handle_delete()
            return
        if route == "/api/data/cache":
            self._handle_cache_post()
            return
        if route == "/api/export/defect-rate-xlsx":
            self._handle_export_defect_rate_xlsx()
            return
        if route == "/api/public-url":
            self._handle_public_url_post()
            return
        self._send_json(404, {"ok": False, "error": "API route not found."})

    def _maybe_consume_token_query(self) -> bool:
        """URL의 ?token= 을 쿠키로 옮기고 깔끔한 주소로 리다이렉트한다."""
        if not self._required_access_token():
            return False
        parsed = urlparse(self.path)
        qs = parse_qs(parsed.query)
        q_vals = qs.get("token") or qs.get("access_token") or []
        if not q_vals or not str(q_vals[0]).strip():
            return False
        clean = parsed.path or "/index.html"
        if clean == "/":
            clean = "/index.html"
        self.send_response(302)
        self._set_access_cookie_header()
        self.send_header("Location", clean)
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        return True

    def _handle_server_info(self) -> None:
        lan_ip = get_lan_ip()
        port = self.server.server_address[1]
        lan_url = f"http://{lan_ip}:{port}/index.html" if lan_ip else ""
        public_url = read_public_tunnel_url(self.server.public_url_path)
        env_public = str(os.environ.get("NCP_PUBLIC_URL") or "").strip().rstrip("/")
        if env_public.startswith("http"):
            public_url = env_public
        token_required = bool(self._required_access_token())
        share_url = ""
        if public_url:
            share_url = f"{public_url}/index.html"
            if token_required:
                share_url = f"{share_url}?token={quote(self._required_access_token(), safe='')}"
        self._send_json(
            200,
            {
                "lanUrl": lan_url,
                "port": port,
                "publicUrl": public_url,
                "shareUrl": share_url,
                "tokenRequired": token_required,
                "anywhereMode": bool(public_url),
            },
        )

    def _handle_public_url_post(self) -> None:
        """터널 스크립트가 공개 HTTPS URL을 등록한다."""
        try:
            payload = self._read_json_body()
            url = str(payload.get("url") or "").strip().rstrip("/")
            if not (url.startswith("http://") or url.startswith("https://")):
                raise ValueError("url must start with http:// or https://")
            path = self.server.public_url_path
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text(
                f"# auto-updated by mobile-anywhere tunnel\n{url}\n",
                encoding="utf-8",
            )
            self._send_json(200, {"ok": True, "publicUrl": url})
        except ValueError as exc:
            self._send_json(400, {"ok": False, "error": str(exc)})
        except Exception:
            self._send_json(500, {"ok": False, "error": "Failed to save public URL."})

    def _handle_manifest_get(self) -> None:
        manifest = load_manifest(self.server.manifest_path)
        self._send_json(200, manifest)

    def _handle_cache_get(self) -> None:
        try:
            cached = load_cache_record(self.server.cache_path)
            if not cached:
                self._send_json(404, {"ok": False, "error": "Cache not found."})
                return
            self._send_json(200, cached)
        except Exception:
            self._send_json(500, {"ok": False, "error": "Failed to read cache file."})

    def _handle_cache_post(self) -> None:
        try:
            length_header = self.headers.get("Content-Length", "")
            try:
                length = int(length_header)
            except ValueError:
                raise ValueError("Invalid Content-Length header.")
            if length <= 0:
                raise ValueError("Request body is empty.")
            if length > MAX_CACHE_BYTES:
                raise ValueError("Cache body is too large.")
            raw = self.rfile.read(length)
            try:
                payload = json.loads(raw.decode("utf-8"))
            except Exception:
                raise ValueError("Invalid JSON body.")
            if not isinstance(payload, dict):
                raise ValueError("JSON body must be an object.")
            record_src = payload.get("record") if isinstance(payload.get("record"), dict) else payload
            normalized = normalize_cache_record(record_src)
            if not normalized:
                raise ValueError("Cache payload must include snRows or swRows.")
            expected_raw = payload.get("expectedVersion")
            expected_version = None
            if expected_raw is not None:
                try:
                    expected_version = int(expected_raw)
                except Exception:
                    raise ValueError("expectedVersion must be an integer.")
            force = bool(payload.get("force", False))
            with self.server.cache_lock:
                # 서버 파일 기준 현재 버전을 확인하고 낙관적 동시성 제어 수행
                current = load_cache_record(self.server.cache_path)
                current_version = to_int(current.get("version"), 0) if current else 0
                if not force and expected_version is not None and expected_version != current_version:
                    self._send_json(409, {
                        "ok": False,
                        "error": "Version conflict.",
                        "currentVersion": current_version,
                    })
                    return

                now_ms = int(datetime.now(timezone.utc).timestamp() * 1000)
                next_version = current_version + 1
                normalized["version"] = next_version
                normalized["updatedAt"] = max(now_ms, to_int(normalized.get("updatedAt"), 0))
                source = normalized.get("source")
                source_obj = source if isinstance(source, dict) else {}
                source_obj["kind"] = str(source_obj.get("kind") or "uploaded_excel_synced")
                if source_obj["kind"] in ("uploaded_excel", "uploaded_excel_cached"):
                    source_obj["kind"] = "uploaded_excel_synced"
                source_obj["label"] = str(source_obj.get("label") or "Shared LAN dataset")
                source_obj["savedAt"] = utc_now_iso()
                normalized["source"] = source_obj
                normalized["fingerprint"] = str(normalized.get("fingerprint") or "").strip() or make_cache_fingerprint(normalized)
                save_cache_record(self.server.cache_path, normalized)
                result = {
                    "ok": True,
                    "version": normalized["version"],
                    "updatedAt": normalized["updatedAt"],
                    "savedAt": source_obj.get("savedAt", ""),
                    "fingerprint": normalized["fingerprint"],
                }
            self._send_json(200, result)
        except ValueError as exc:
            self._send_json(400, {"ok": False, "error": str(exc)})
        except Exception:
            self._send_json(500, {"ok": False, "error": "Internal error while saving cache file."})

    def _handle_upload(self) -> None:
        try:
            payload = self._read_json_body()
            row_key = str(payload.get("rowKey", "")).strip()
            if not row_key:
                raise ValueError("rowKey is required.")
            if len(row_key) > 300:
                raise ValueError("rowKey is too long.")

            name = str(payload.get("name", "")).strip() or "attachment_file"
            file_type = str(payload.get("type", "")).strip() or "application/octet-stream"
            ext = Path(name).suffix.lower()
            if ext not in ALLOWED_ATTACH_EXTENSIONS:
                allowed = ", ".join(sorted(ALLOWED_ATTACH_EXTENSIONS))
                raise ValueError(f"Unsupported file extension: {ext or '(none)'} (allowed: {allowed})")
            data_base64 = payload.get("dataBase64")
            if not isinstance(data_base64, str) or not data_base64:
                raise ValueError("dataBase64 is required.")

            try:
                file_bytes = base64.b64decode(data_base64, validate=True)
            except Exception:
                raise ValueError("Failed to decode dataBase64.")
            if not file_bytes:
                raise ValueError("File data is empty.")
            if len(file_bytes) > MAX_UPLOAD_BYTES:
                raise ValueError("Upload size exceeds 30MB limit.")
            with self.server.manifest_lock:
                # 첨부 파일 저장과 manifest 갱신을 한 구간에서 처리해 일관성 유지
                self.server.attach_dir.mkdir(parents=True, exist_ok=True)
                filename = make_storage_filename(row_key, name)
                target = self.server.attach_dir / filename
                target.write_bytes(file_bytes)
                now_ms = int(datetime.now(timezone.utc).timestamp() * 1000)
                manifest = load_manifest(self.server.manifest_path)
                previous = manifest.get("items", {}).get(row_key)
                item = {
                    "name": name,
                    "type": file_type,
                    "size": len(file_bytes),
                    "updatedAt": now_ms,
                    "path": f"cloud/sw_attachments/{filename}",
                }
                manifest["version"] = 1
                manifest["updatedAt"] = utc_now_iso()
                manifest["items"][row_key] = item
                save_manifest(self.server.manifest_path, manifest)
                if previous:
                    # 같은 rowKey 재업로드(교체) 시 이전 파일 정리
                    rel_path = str(previous.get("path", "")).replace("\\", "/")
                    old_name = Path(rel_path).name
                    if old_name and old_name != filename:
                        old_target = self.server.attach_dir / old_name
                        try:
                            if old_target.exists() and old_target.is_file():
                                old_target.unlink()
                        except OSError:
                            pass
            self._send_json(200, {"ok": True, "item": item})
        except ValueError as exc:
            self._send_json(400, {"ok": False, "error": str(exc)})
        except Exception:
            self._send_json(500, {"ok": False, "error": "Internal error while saving shared attachment."})

    def _handle_delete(self) -> None:
        try:
            payload = self._read_json_body()
            row_key = str(payload.get("rowKey", "")).strip()
            if not row_key:
                raise ValueError("rowKey is required.")
            with self.server.manifest_lock:
                # manifest에서 제거하고 실제 파일도 같이 정리
                manifest = load_manifest(self.server.manifest_path)
                items = manifest.get("items", {})
                item = items.pop(row_key, None)
                deleted = item is not None
                if item:
                    rel_path = str(item.get("path", "")).replace("\\", "/")
                    file_name = Path(rel_path).name
                    if file_name:
                        target = self.server.attach_dir / file_name
                        try:
                            if target.exists() and target.is_file():
                                target.unlink()
                        except OSError:
                            pass
                manifest["version"] = 1
                manifest["updatedAt"] = utc_now_iso()
                manifest["items"] = items
                save_manifest(self.server.manifest_path, manifest)
            self._send_json(200, {"ok": True, "deleted": deleted})
        except ValueError as exc:
            self._send_json(400, {"ok": False, "error": str(exc)})
        except Exception:
            self._send_json(500, {"ok": False, "error": "Internal error while deleting shared attachment."})

    def _handle_export_defect_rate_xlsx(self) -> None:
        try:
            payload = self._read_json_body()
            file_bytes, file_name = _build_defect_rate_export_xlsx(payload)
            self._send_binary(
                200,
                file_bytes,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                download_name=file_name,
            )
        except ValueError as exc:
            self._send_json(400, {"ok": False, "error": str(exc)})
        except RuntimeError as exc:
            self._send_json(500, {"ok": False, "error": str(exc)})
        except Exception:
            self._send_json(500, {"ok": False, "error": "Internal error while building export workbook."})


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Local static server with shared attachment API")
    parser.add_argument("--host", default="0.0.0.0", help="Bind host (default: 0.0.0.0)")
    parser.add_argument("--port", type=int, default=8787, help="Bind port (default: 8787)")
    parser.add_argument("--root", default=str(Path(__file__).resolve().parents[1]), help="Repository root path")
    parser.add_argument(
        "--allow-public-clients",
        action="store_true",
        # 기본은 내부망만 허용, 필요 시에만 공개 허용
        help="Allow public/global client IP addresses (default: private/LAN only).",
    )
    return parser.parse_args()


def _is_address_in_use(exc: BaseException) -> bool:
    if not isinstance(exc, OSError):
        return False
    if getattr(exc, "winerror", None) == 10048:
        return True
    return exc.errno == errno.EADDRINUSE


def main() -> int:
    args = parse_args()
    repo_root = Path(args.root).resolve()
    if not repo_root.exists():
        print(f"Root path not found: {repo_root}", file=sys.stderr)
        return 1
    if args.port < 1 or args.port > 65535:
        print("Port must be between 1 and 65535.", file=sys.stderr)
        return 1

    os.chdir(repo_root)
    handler = partial(SharedAttachmentHandler, directory=str(repo_root))

    bind_port = args.port
    server: SharedAttachmentServer | None = None
    max_tries = 32
    for _ in range(max_tries):
        try:
            server = SharedAttachmentServer(
                (args.host, bind_port),
                handler,
                repo_root,
                allow_public_clients=args.allow_public_clients,
            )
            break
        except OSError as e:
            if _is_address_in_use(e) and bind_port < 65535:
                if bind_port == args.port:
                    print(
                        f"[WARN] 포트 {args.port} 은(는) 이미 사용 중입니다. "
                        f"다른 포트를 찾는 중… (WinError 10048 / EADDRINUSE)",
                        file=sys.stderr,
                    )
                bind_port += 1
                continue
            print(f"[ERROR] 서버를 시작할 수 없습니다: {e}", file=sys.stderr)
            return 1

    if server is None:
        print("[ERROR] 사용 가능한 포트를 찾지 못했습니다.", file=sys.stderr)
        return 1

    if bind_port != args.port:
        print(
            f"[INFO] 실제 접속 포트: {bind_port} "
            f"(요청 {args.port} 대체). 방화벽은 해당 포트를 허용해야 할 수 있습니다.",
            file=sys.stderr,
        )

    lan_ip = get_lan_ip()
    lan_url = f"http://{lan_ip}:{bind_port}/index.html" if lan_ip else ""
    print("=" * 60)
    print("  Shinsung QC Report - Local Server")
    print("=" * 60)
    print(f"  Root : {repo_root}")
    print(f"  Local: http://127.0.0.1:{bind_port}/index.html")
    if lan_url and args.host in ("0.0.0.0", ""):
        print(f"  LAN  : {lan_url}  (use this URL on mobile/other PCs)")
    print("-" * 60)
    print("  API: /api/sw-attachments/manifest  (list attachments)")
    print("  API: /api/sw-attachments/upload    (upload attachment)")
    print("  API: /api/sw-attachments/delete    (delete attachment)")
    print("  API: /api/data/cache               (get/save report cache)")
    print("  API: /api/export/defect-rate-xlsx (export workbook with chart)")
    print("  Static gzip: enabled for text assets")
    print("-" * 60)
    print("  Stop : Ctrl+C")
    policy = "allow all clients" if args.allow_public_clients else "private/LAN clients only"
    print(f"  Access policy: {policy}")
    print("=" * 60)
    if lan_url and args.host in ("0.0.0.0", ""):
        print_qr_code(lan_url)

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
